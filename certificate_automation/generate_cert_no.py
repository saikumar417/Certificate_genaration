import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import os
import traceback
import datetime


def generate_sequence(program_count, start, num_records, year):
    sequence = []
    for i in range(start, start + num_records):
        sequence.append(f"SAA/{program_count}/{str(i).zfill(4)}/{year}")
    return sequence


def log_error(error_message):
    """Log errors to a file."""
    with open("error_log.txt", "a") as log_file:
        log_file.write(f"{datetime.datetime.now()} - {error_message}\n")


class CertificateNumberApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Certificate Sequence Generator")
        self.root.geometry("500x400")
        self.root.configure(bg="#f0f8ff")

        self.excel_file_path = ""  # To store the selected Excel file path

        # Title Label
        title_label = tk.Label(
            root, text="Certificate Sequence Generator", font=("Helvetica", 16, "bold"), bg="#f0f8ff", fg="#00008b"
        )
        title_label.pack(pady=10)

        # Program Count Input
        tk.Label(root, text="Program Count (e.g., 01):", bg="#f0f8ff", fg="#00008b").pack(anchor="w", padx=20)
        self.program_input = tk.Entry(root, font=("Helvetica", 12), bg="#ffffff")
        self.program_input.pack(fill="x", padx=20, pady=5)

        # Start Number Input
        tk.Label(root, text="Start Number (e.g., 1 for 0001):", bg="#f0f8ff", fg="#00008b").pack(anchor="w", padx=20)
        self.start_input = tk.Entry(root, font=("Helvetica", 12), bg="#ffffff")
        self.start_input.pack(fill="x", padx=20, pady=5)

        # Year Input
        tk.Label(root, text="Year (e.g., 2024):", bg="#f0f8ff", fg="#00008b").pack(anchor="w", padx=20)
        self.year_input = tk.Entry(root, font=("Helvetica", 12), bg="#ffffff")
        self.year_input.insert(0, str(datetime.datetime.now().year))  # Default to the current year
        self.year_input.pack(fill="x", padx=20, pady=5)

        # File Selection Section
        self.file_label = tk.Label(
            root, text="No file selected", font=("Helvetica", 10), bg="#f0f8ff", fg="#ff4500"
        )
        self.file_label.pack(pady=5)

        file_buttons_frame = tk.Frame(root, bg="#f0f8ff")
        file_buttons_frame.pack(pady=5)

        upload_button = tk.Button(
            file_buttons_frame, text="Upload Excel File", command=self.choose_file, bg="#4682b4", fg="#ffffff", font=("Helvetica", 10, "bold")
        )
        upload_button.pack(side="left", padx=10)

        # Generate Button
        generate_button = tk.Button(
            root, text="Generate Sequence", command=self.generate_file, bg="#228b22", fg="#ffffff", font=("Helvetica", 12, "bold")
        )
        generate_button.pack(pady=10)

        # Help Button
        help_button = tk.Button(
            root, text="Help", command=self.show_help, bg="#ff8c00", fg="#ffffff", font=("Helvetica", 10, "bold")
        )
        help_button.pack(pady=10)

    def choose_file(self):
        try:
            file_path = filedialog.askopenfilename(
                title="Select an Excel File",
                filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            )
            if file_path:
                self.excel_file_path = file_path
                file_size = os.path.getsize(file_path) / (1024 * 1024)  # Convert bytes to MB
                self.file_label.config(
                    text=f"Selected file: {os.path.basename(file_path)} ({file_size:.2f} MB)", fg="#228b22"
                )
            else:
                self.file_label.config(text="No file selected", fg="#ff4500")
        except Exception as e:
            error_message = traceback.format_exc()
            log_error(error_message)
            messagebox.showerror("Error", f"Failed to select file:\n{e}")

    def generate_file(self):
        try:
            program_count = self.program_input.get() or "01"
            start = self.start_input.get()
            year = self.year_input.get()

            if not start.isdigit() or int(start) < 1:
                messagebox.showerror("Error", "Start Number must be a positive integer.")
                return

            if not year.isdigit() or len(year) != 4:
                messagebox.showerror("Error", "Year must be a 4-digit number.")
                return

            if not self.excel_file_path:
                messagebox.showerror("Error", "Please upload an Excel file.")
                return

            # Load Excel File
            wb = openpyxl.load_workbook(self.excel_file_path)
            ws = wb.active

            # Insert sequence column
            if ws["A1"].value != "Certno":
                ws.insert_cols(1)
                ws["A1"] = "Certno"

            # Generate sequences
            start_num = int(start)
            num_records = sum(1 for row in ws.iter_rows(min_row=2) if row[1].value)
            sequences = generate_sequence(program_count, start_num, num_records, year)

            for i, seq in enumerate(sequences, start=2):
                ws[f"A{i}"] = seq

            # Save updated file (overwriting the original file)
            wb.save(self.excel_file_path)
            wb.close()

            messagebox.showinfo("Success", f"Sequence added and saved in:\n{self.excel_file_path}")

        except Exception as e:
            error_message = traceback.format_exc()
            log_error(error_message)
            messagebox.showerror("Error", f"An unexpected error occurred:\n{e}")

    def show_help(self):
        """Display a help popup."""
        help_message = (
            "Instructions:\n"
            "- Enter the Program Count (e.g., 01).\n"
            "- Enter the Start Number (e.g., 1).\n"
            "- Enter the Year (defaults to the current year).\n"
            "- Upload an Excel file (.xlsx) with data starting from the second row.\n"
            "- Click 'Generate Sequence' to update the file in-place."
        )
        messagebox.showinfo("Help", help_message)


def main():
    root = tk.Tk()
    app = CertificateNumberApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()

